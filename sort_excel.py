#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sort_excel.py (improved)
- Excelの「受諾確認票」シートを読み込み、3種類のシートを生成
  1) 融資実行日でソート
  2) 金消日・面談日でソート
  3) 上記を統合し、日付ごとの個別シートも自動作成
- 「日付シート」では、以下の優先順位で並べ替え
   (1) 識別: 降順
   (2) 担当: 指定の順序（未リスト名は最後）
   (3) 時間: 昇順
- 罫線/高さ/配置/幅/縮小表示など表の書式を自動設定
- PyInstallerでexe化できるよう、余計な依存を避けた実装
"""

import sys
from pathlib import Path
import pandas as pd
import numpy as np

# ====== 設定 ======
担当順 = ['河内','岩川','杉田','正木','吉田','脇本','中本','椙村','北条','上野']

# Excelの列見出し（標準化後）
COLS_LOAN   = ['融資実行日','形態','お客様氏名','物件','依頼内容','担当','管轄','立会時間','立会場所','立会者','当日申請']
COLS_CANCEL = ['金消日・面談日','形態','お客様氏名','物件','依頼内容','担当','管轄','金消時間','金消場所・面談場所','意思確認','融資実行日']

COLS_OUT_ORDER = ['日付','形態','お客様氏名','物件','依頼内容','担当','管轄','時間','場所','確認者','申請','識別']

# 書式設定（列幅はだいたいのピクセル→Excel幅換算（約 1単位≒7px））
PX = lambda p: round(p/7.0, 2)
WIDTH_RULES = {
    'お客様氏名': PX(115),
    '物件': PX(160),
    '場所': PX(160),
    # その他は 80px
}
DEFAULT_WIDTH = PX(80)

ROW_HEIGHT = 25  # pt ではなく「行の高さ」単位（ExcelのUI上）
# ==================

def _read_sheet(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Sheet1')
    df.columns = df.columns.str.strip()
    # 日付系
    for c in ['融資実行日','金消日・面談日']:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors='coerce')
    return df

def _dt_to_str(s: pd.Series) -> pd.Series:
    # NaTはそのまま
    return s.dt.strftime('%Y/%m/%d')

def _make_tables(df: pd.DataFrame):
    # (1) 融資実行日でソート
    loan = df.copy()
    loan = loan.sort_values(by='融資実行日')
    loan = loan.assign(
        識別 = '融資実行日'
    )
    loan = loan.rename(columns={'融資実行日':'日付','立会時間':'時間','立会場所':'場所','立会者':'確認者','当日申請':'申請'})
    # 列整列
    loan = loan[[c for c in COLS_OUT_ORDER if c in loan.columns]]
    # 日付文字列化
    if '日付' in loan.columns:
        loan['日付'] = pd.to_datetime(loan['日付'], errors='coerce')
        loan['日付'] = _dt_to_str(loan['日付'])

    # (2) 金消日・面談日でソート
    cancel = df.copy()
    cancel = cancel.sort_values(by='金消日・面談日')
    cancel = cancel.assign(
        識別 = '金消日・面談日'
    )
    cancel = cancel.rename(columns={'金消日・面談日':'日付','金消時間':'時間','金消場所・面談場所':'場所','意思確認':'確認者','融資実行日':'申請'})
    cancel = cancel[[c for c in COLS_OUT_ORDER if c in cancel.columns]]
    if '日付' in cancel.columns:
        cancel['日付'] = pd.to_datetime(cancel['日付'], errors='coerce')
        cancel['日付'] = _dt_to_str(cancel['日付'])

    # (3) 統合 + 「日付」昇順
    combined = pd.concat([loan, cancel], ignore_index=True, sort=False)
    # 「日付」昇順（空は最後）
    tmp = pd.to_datetime(combined['日付'], errors='coerce')
    combined = combined.assign(_d=tmp).sort_values(by=['_d']).drop(columns=['_d'])

    # 出力列の最終順
    combined = combined.reindex(columns=[c for c in COLS_OUT_ORDER if c in combined.columns])

    return loan, cancel, combined

def _parse_time_like(s: pd.Series) -> pd.Series:
    """
    'HH:MM' 'H:MM' 'HH:MM:SS' などの時間文字列を秒へ。
    数値やNaNは安全にNaNへ。
    """
    if s is None:
        return pd.Series(dtype='float64')
    ss = s.fillna('')
    def tosec(x):
        t = str(x).strip()
        if not t:
            return np.nan
        # 9:30 / 09:30 / 09:30:15 など
        parts = t.split(':')
        try:
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            sec = int(parts[2]) if len(parts) > 2 else 0
            return h*3600 + m*60 + sec
        except Exception:
            return np.nan
    return ss.map(tosec).astype('float64')

def _sort_for_date_sheet(df_date: pd.DataFrame) -> pd.DataFrame:
    """
    日付シートの並べ替え規則:
      1) 識別: 降順
      2) 担当: 指定順（未掲載は最後）
      3) 時間: 昇順
    """
    # 識別: 文字列の降順（必要あればカスタム順に変更可）
    key1 = df_date['識別'].astype(str)

    # 担当: カスタム順（未リストは大きな値）
    order = {name: i for i, name in enumerate(担当順)}
    key2 = df_date['担当'].astype(str).map(order).fillna(9999).astype(int)

    # 時間: 可能な限り数値（秒）にパース
    key3 = _parse_time_like(df_date['時間'])

    df_sorted = df_date.assign(_k1=key1, _k2=key2, _k3=key3)\
                       .sort_values(by=['_k1','_k2','_k3'], ascending=[True, True, True])\
                       .drop(columns=['_k1','_k2','_k3'])
    return df_sorted

def _write_excel(out_path: Path, loan: pd.DataFrame, cancel: pd.DataFrame, combined: pd.DataFrame):
    # まずpandasで出力
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        loan.to_excel(writer, sheet_name='sorted_by_融資実行日', index=False)
        cancel.to_excel(writer, sheet_name='sorted_by_金消日・面談日', index=False)
        combined.to_excel(writer, sheet_name='統合データ', index=False)

        # 日付ごとのシート
        if '日付' in combined.columns:
            for date in combined['日付'].dropna().unique():
                df_date = combined[combined['日付'] == date].copy()
                df_date = _sort_for_date_sheet(df_date)
                sheet = str(date).replace('/','-')
                df_date.to_excel(writer, sheet_name=sheet, index=False)

    # ここから openpyxl で書式を整える
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font

    wb = load_workbook(out_path)

    # 共通スタイル
    dotted = Side(style='dotted')
    border_all = Border(left=dotted, right=dotted, top=dotted, bottom=dotted)
    align = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrap_text=False)
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        # 全体セル範囲
        max_row = ws.max_row
        max_col = ws.max_column

        # ヘッダ行を少し強調 & 高さ
        ws.row_dimensions[1].height = ROW_HEIGHT

        # 罫線・配置・高さ
        for r in range(1, max_row+1):
            ws.row_dimensions[r].height = ROW_HEIGHT
            for c in range(1, max_col+1):
                cell = ws.cell(row=r, column=c)
                cell.border = border_all
                # ヘッダは太字
                if r == 1:
                    cell.font = header_font
                    # ヘッダは中央にしつつ shrink_to_fit
                    cell.alignment = align
                else:
                    cell.alignment = align

        # 列幅設定
        header_names = [ws.cell(row=1, column=c).value for c in range(1, max_col+1)]
        for idx, name in enumerate(header_names, start=1):
            if name in WIDTH_RULES:
                width = WIDTH_RULES[name]
            else:
                width = DEFAULT_WIDTH
            ws.column_dimensions[chr(64+idx) if idx<=26 else None].width = width  # simple A-Z

        # 万一26列を超える場合の安全対策
        for c in range(1, max_col+1):
            col_letter = ws.cell(row=1, column=c).column_letter
            name = ws.cell(row=1, column=c).value
            width = WIDTH_RULES.get(name, DEFAULT_WIDTH)
            ws.column_dimensions[col_letter].width = width

    wb.save(out_path)

def sort_excel(input_path: str, output_path: str = 'sorted_combined.xlsx'):
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f'入力ファイルが見つかりません: {src}')
    df = _read_sheet(src)
    loan, cancel, combined = _make_tables(df)
    _write_excel(Path(output_path), loan, cancel, combined)
    return str(Path(output_path).resolve())

def main():
    if len(sys.argv) < 2:
        print("使い方: sort_excel.exe <Excelファイルパス>  あるいは  python sort_excel.py <Excelファイルパス>")
        sys.exit(1)
    input_file = sys.argv[1]
    out = sort_excel(input_file)
    print(f'出力: {out}')

if __name__ == '__main__':
    main()
